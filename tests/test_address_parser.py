from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
import pandas as pd
import pytest

from address_parser import parse_address
from data_processing import (
    create_excel_workbook,
    detect_columns,
    process_dataframe,
    read_csv_bytes,
)


def assert_parsed(
    source: str,
    *,
    address: str,
    suite: str = "",
    city: str,
    zip_code: str,
) -> None:
    result = parse_address(source)
    assert result["Address"] == address
    assert result["Suite"] == suite
    assert result["City"] == city
    assert result["Zip Code"] == zip_code
    assert result["Parse Status"] == "Parsed"


def test_sample_unlabeled_numeric_suite() -> None:
    assert_parsed(
        "7744 BROADWAY ST 203 SAN ANTONIO, TX 78216",
        address="7744 BROADWAY ST",
        suite="203",
        city="SAN ANTONIO",
        zip_code="78216",
    )


def test_sample_normal_address_without_suite() -> None:
    assert_parsed(
        "910 BROADWAY ST SAN ANTONIO, TX 78215",
        address="910 BROADWAY ST",
        city="SAN ANTONIO",
        zip_code="78215",
    )


def test_sample_zip_plus_four() -> None:
    assert_parsed(
        "120 BROADWAY ST SAN ANTONIO, TX 78205-1904",
        address="120 BROADWAY ST",
        city="SAN ANTONIO",
        zip_code="78205-1904",
    )


def test_sample_second_unlabeled_suite_and_zip_plus_four() -> None:
    assert_parsed(
        "125 BROADWAY ST 300 SAN ANTONIO, TX 78205-1903",
        address="125 BROADWAY ST",
        suite="300",
        city="SAN ANTONIO",
        zip_code="78205-1903",
    )


@pytest.mark.parametrize(
    ("label_and_suite", "expected_suite"),
    [
        ("SUITE 200", "200"),
        ("STE 200A", "200A"),
        ("UNIT A-12", "A-12"),
        ("APT 4B", "4B"),
        ("#444", "444"),
        ("BUILDING C", "C"),
        ("BLDG 7", "7"),
        ("FLOOR 4", "4"),
        ("FL 2", "2"),
    ],
)
def test_explicit_suite_labels(label_and_suite: str, expected_suite: str) -> None:
    result = parse_address(f"500 Main Street {label_and_suite} Austin, TX 78701")
    assert result["Parse Status"] == "Parsed"
    assert result["Address"] == "500 Main Street"
    assert result["Suite"] == expected_suite
    assert result["City"] == "Austin"


def test_numbered_street_is_not_a_suite() -> None:
    assert_parsed(
        "311 10TH ST SAN ANTONIO, TX 78215",
        address="311 10TH ST",
        city="SAN ANTONIO",
        zip_code="78215",
    )


def test_directional_street_component_is_retained() -> None:
    assert_parsed(
        "927 N ALAMO ST SAN ANTONIO, TX 78215",
        address="927 N ALAMO ST",
        city="SAN ANTONIO",
        zip_code="78215",
    )


def test_multiword_street() -> None:
    assert_parsed(
        "42 OLD MILL ROAD BOSTON, MA 02108",
        address="42 OLD MILL ROAD",
        city="BOSTON",
        zip_code="02108",
    )


def test_multiword_city_in_another_state() -> None:
    assert_parsed(
        "1600 PENNSYLVANIA AVENUE WASHINGTON, DC 20500",
        address="1600 PENNSYLVANIA AVENUE",
        city="WASHINGTON",
        zip_code="20500",
    )


def test_city_beginning_with_st_is_not_mistaken_for_street_suffix() -> None:
    assert_parsed(
        "123 MARKET ST ST LOUIS, MO 63101",
        address="123 MARKET ST",
        city="ST LOUIS",
        zip_code="63101",
    )


def test_missing_commas() -> None:
    assert_parsed(
        "100 Congress Ave Austin TX 78701",
        address="100 Congress Ave",
        city="Austin",
        zip_code="78701",
    )


def test_extra_spaces_are_normalized() -> None:
    assert_parsed(
        "  100   Congress   Ave   Suite  9   Austin,   TX   78701  ",
        address="100 Congress Ave",
        suite="9",
        city="Austin",
        zip_code="78701",
    )


def test_mixed_capitalization_is_preserved_and_parsed() -> None:
    assert_parsed(
        "100 cOnGrEsS aVe aUsTiN, tX 78701",
        address="100 cOnGrEsS aVe",
        city="aUsTiN",
        zip_code="78701",
    )


@pytest.mark.parametrize("blank", ["", "   ", None, float("nan")])
def test_blank_address_requires_review(blank) -> None:
    result = parse_address(blank)
    assert result["Parse Status"] == "Review Needed"
    assert result["Address"] == ""
    assert "blank" in result["Parse Notes"].lower()


def test_malformed_address_requires_review_without_discarding_text() -> None:
    result = parse_address("NOT AN ADDRESS")
    assert result["Parse Status"] == "Review Needed"
    assert result["Address"] == "NOT AN ADDRESS"


def test_ambiguous_unlabeled_number_is_kept_for_review() -> None:
    result = parse_address("123 MAIN ST 12 34 AUSTIN, TX 78701")
    assert result["Parse Status"] == "Review Needed"
    assert result["Suite"] == ""
    assert "12 34" in result["Address"]
    assert "number" in result["Parse Notes"].lower()


def test_unlabeled_number_without_state_is_not_guessed_as_suite() -> None:
    result = parse_address("123 MAIN ST 200 AUSTIN 78701")
    assert result["Parse Status"] == "Review Needed"
    assert result["Suite"] == ""
    assert "200" in result["Address"]


def test_missing_zip_is_partial_when_other_components_parse() -> None:
    result = parse_address("123 MAIN ST AUSTIN, TX")
    assert result["Parse Status"] == "Partial"
    assert result["Address"] == "123 MAIN ST"
    assert result["City"] == "AUSTIN"
    assert result["Zip Code"] == ""


def test_missing_state_is_partial_when_other_components_parse() -> None:
    result = parse_address("123 MAIN ST AUSTIN 78701")
    assert result["Parse Status"] == "Partial"
    assert result["City"] == "AUSTIN"


@pytest.mark.parametrize(
    "source",
    [
        "1 OAK BLVD DALLAS, TX 75001",
        "2 PINE DR DENVER, CO 80202",
        "3 ELM LN PORTLAND, OR 97201",
        "4 CEDAR PKWY PHOENIX, AZ 85001",
        "5 MAPLE CT MIAMI, FL 33101",
        "6 BIRCH CIR CHICAGO, IL 60601",
        "7 ASPEN PL BOISE, ID 83702",
        "8 WALNUT TER ATLANTA, GA 30301",
        "9 SPRUCE TRL NASHVILLE, TN 37201",
    ],
)
def test_supported_street_suffixes(source: str) -> None:
    assert parse_address(source)["Parse Status"] == "Parsed"


def make_source_frame() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Property ID": ["001", "002", "003", "004", "005"],
            "Geographic ID": ["000-A", "000-B", "000-C", "000-D", "000-E"],
            "Type": ["Personal", " REAL ", " personal ", "", "Commercial"],
            "PropertyAddress": [
                "910 BROADWAY ST SAN ANTONIO, TX 78215",
                "1 MAIN ST AUSTIN, TX 78701",
                "125 BROADWAY ST 300 SAN ANTONIO, TX 78205-1903",
                "2 MAIN ST AUSTIN, TX 78701",
                "3 MAIN ST AUSTIN, TX 78701",
            ],
            " legal description ": ["remove me"] * 5,
            "Owner Name": ["A", "B", "C", "D", "E"],
            "Custom Field": ["keep-1", "keep-2", "keep-3", "keep-4", "keep-5"],
        }
    )


def test_processing_filters_type_case_insensitively_and_counts_exclusions() -> None:
    result = process_dataframe(make_source_frame())
    assert result.database["Property ID"].tolist() == ["001", "003"]
    assert result.summary["personal_retained"] == 2
    assert result.summary["real_removed"] == 1
    assert result.summary["blank_or_unexpected_excluded"] == 2


def test_processing_removes_legal_description_case_insensitively() -> None:
    result = process_dataframe(make_source_frame())
    assert "Legal Description" not in result.database.columns
    assert " legal description " not in result.database.columns


def test_processing_preserves_original_address_and_adds_parsed_columns() -> None:
    source = make_source_frame()
    original_address = source.loc[0, "PropertyAddress"]
    result = process_dataframe(source)
    assert result.database.loc[0, "PropertyAddress"] == original_address
    for column in ("Address", "Suite", "City", "Zip Code", "Parse Status", "Parse Notes"):
        assert column in result.database.columns


def test_processing_preserves_extra_columns_after_primary_columns() -> None:
    result = process_dataframe(make_source_frame())
    assert result.database["Custom Field"].tolist() == ["keep-1", "keep-3"]
    assert result.database.columns[-1] == "Custom Field"


def test_processing_preserves_input_columns_that_collide_with_generated_names() -> None:
    source = make_source_frame().assign(Address="legacy address value")
    result = process_dataframe(source)
    assert result.database["Original Address"].tolist() == [
        "legacy address value",
        "legacy address value",
    ]
    assert result.database.loc[0, "Address"] == "910 BROADWAY ST"


def test_processing_handles_zero_personal_rows() -> None:
    source = make_source_frame().assign(Type="Real")
    result = process_dataframe(source)
    assert result.database.empty
    assert result.review_needed.empty
    assert result.summary["personal_retained"] == 0


def test_header_detection_ignores_case_spaces_and_minor_variations() -> None:
    detection = detect_columns([" TYPE ", " property Address ", "Legal description"])
    assert detection.errors == []
    assert detection.type_column == " TYPE "
    assert detection.property_address_column == " property Address "
    assert detection.legal_description_column == "Legal description"


def test_header_detection_reports_missing_required_columns() -> None:
    detection = detect_columns(["Owner Name", "Legal Description"])
    assert len(detection.errors) == 2


def test_utf8_bom_csv_and_leading_zero_identifiers_are_preserved() -> None:
    raw = (
        "Property ID,Type,PropertyAddress\n"
        '"00123",Personal,"1 MAIN ST AUSTIN, TX 78701"\n'
    ).encode("utf-8-sig")
    frame, encoding = read_csv_bytes(raw)
    assert encoding == "utf-8-sig"
    assert frame.loc[0, "Property ID"] == "00123"


def test_latin1_csv_fallback() -> None:
    raw = (
        "Type,PropertyAddress,Owner Name\n"
        'Personal,"1 MAIN ST AUSTIN, TX 78701",Jos\xe9\n'
    ).encode("latin-1")
    frame, encoding = read_csv_bytes(raw)
    assert encoding == "latin-1"
    assert frame.loc[0, "Owner Name"] == "José"


def test_excel_workbook_has_required_sheets_formatting_and_text_ids() -> None:
    result = process_dataframe(make_source_frame())
    workbook_bytes = create_excel_workbook(result)
    workbook = load_workbook(BytesIO(workbook_bytes))
    assert workbook.sheetnames == ["Database", "Review Needed", "Processing Summary"]
    database = workbook["Database"]
    assert database.freeze_panes == "A2"
    assert database.auto_filter.ref
    assert database["A2"].value == "001"
    assert database["A2"].number_format == "@"
    assert database["A1"].font.bold is True


def test_review_sheet_contains_partial_and_review_rows_only() -> None:
    source = pd.DataFrame(
        {
            "Type": ["Personal", "Personal", "Personal"],
            "PropertyAddress": [
                "1 MAIN ST AUSTIN, TX 78701",
                "2 MAIN ST AUSTIN, TX",
                "MALFORMED",
            ],
        }
    )
    result = process_dataframe(source)
    assert set(result.review_needed["Parse Status"]) == {"Partial", "Review Needed"}
    assert len(result.review_needed) == 2
