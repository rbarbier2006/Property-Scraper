from __future__ import annotations

from dataclasses import replace
from io import BytesIO
import json
from types import SimpleNamespace

from openpyxl import load_workbook
import pandas as pd
import pytest

from data_processing import create_excel_workbook, process_dataframe
from openai_reviewer import (
    ALLOWED_REVIEW_FIELDS,
    AddressReviewProposal,
    DEFAULT_OPENAI_MODEL,
    MissingAPIKeyError,
    ReviewServiceError,
    ReviewerConfig,
    build_review_payload,
    is_auto_acceptable,
    load_reviewer_config,
    review_address_with_openai,
    validate_ai_proposal,
)
from review_workflow import (
    apply_review_editor,
    build_review_editor,
    review_eligible_rows,
)


class FakeResponses:
    def __init__(self, output=None, error: Exception | None = None):
        self.output = output
        self.error = error
        self.calls: list[dict] = []

    def parse(self, **kwargs):
        self.calls.append(kwargs)
        if self.error:
            raise self.error
        return SimpleNamespace(output_parsed=self.output)


class FakeClient:
    def __init__(self, output=None, error: Exception | None = None):
        self.responses = FakeResponses(output=output, error=error)


def current_unresolved() -> dict[str, str]:
    return {
        "PropertyAddress": "6104 BROADWAY ST (C1) ALAMO HEIGHTS",
        "Address": "6104 BROADWAY ST (C1) ALAMO HEIGHTS",
        "Suite": "",
        "City": "",
        "Zip Code": "",
        "Parse Status": "Review Needed",
    }


def valid_proposal(**overrides) -> AddressReviewProposal:
    values = {
        "corrected_address": "6104 BROADWAY ST",
        "corrected_suite": "C1",
        "corrected_city": "ALAMO HEIGHTS",
        "corrected_zip_code": "",
        "decision": "Corrected",
        "confidence": 0.98,
        "explanation": "Moved the existing C1 token into Suite.",
        "information_added": False,
        "manual_review_required": False,
    }
    values.update(overrides)
    return AddressReviewProposal(**values)


def reviewer_config(**overrides) -> ReviewerConfig:
    values = {"api_key": "test-placeholder-key", "model": "gpt-5-nano"}
    values.update(overrides)
    return ReviewerConfig(**values)


def unresolved_result():
    return process_dataframe(
        pd.DataFrame(
            {
                "Property ID": ["001", "002"],
                "Type": ["Personal", "Personal"],
                "PropertyAddress": [
                    "1 MAIN ST AUSTIN, TX 78701",
                    "6104 BROADWAY ST (C1) ALAMO HEIGHTS",
                ],
            }
        )
    )


def test_default_model_and_configuration_sources(monkeypatch) -> None:
    for key in (
        "OPENAI_API_KEY",
        "OPENAI_MODEL",
        "OPENAI_MAX_REVIEW_ROWS",
        "OPENAI_AUTO_ACCEPT",
    ):
        monkeypatch.delenv(key, raising=False)
    config = load_reviewer_config({})
    assert config.model == DEFAULT_OPENAI_MODEL == "gpt-5-nano"
    assert config.api_key == ""
    assert config.max_review_rows == 25
    assert config.auto_accept is False

    monkeypatch.setenv("OPENAI_MODEL", "environment-model")
    config = load_reviewer_config(
        {
            "OPENAI_MODEL": "secret-model",
            "OPENAI_MAX_REVIEW_ROWS": 7,
            "OPENAI_AUTO_ACCEPT": True,
        }
    )
    assert config.model == "secret-model"
    assert config.max_review_rows == 7
    assert config.auto_accept is True


def test_payload_contains_only_allow_list_and_rejects_parsed_rows() -> None:
    row = current_unresolved() | {
        "Property ID": "secret-id",
        "Owner Name": "Private Owner",
        "Appraised Value": "999",
    }
    payload = build_review_payload(row, parse_notes="internal")
    assert tuple(payload) == ALLOWED_REVIEW_FIELDS
    assert "Property ID" not in payload
    assert "Owner Name" not in payload
    assert "Appraised Value" not in payload

    row["Parse Status"] = "Parsed"
    with pytest.raises(ValueError):
        build_review_payload(row)


def test_responses_api_uses_pydantic_output_model_no_tools_and_configured_model() -> None:
    proposal = valid_proposal()
    client = FakeClient(output=proposal)
    payload = build_review_payload(current_unresolved())
    returned = review_address_with_openai(payload, reviewer_config(), client=client)
    assert returned == proposal
    call = client.responses.calls[0]
    assert call["model"] == "gpt-5-nano"
    assert call["text_format"] is AddressReviewProposal
    assert call["store"] is False
    assert "tools" not in call
    sent = call["input"][1]["content"]
    sent_json = json.loads(sent.split("\n", 1)[1])
    assert tuple(sent_json) == ALLOWED_REVIEW_FIELDS


def test_valid_ai_correction_passes_python_validation() -> None:
    current = current_unresolved()
    validation = validate_ai_proposal(
        current["PropertyAddress"], current, valid_proposal()
    )
    assert validation.valid is True
    assert is_auto_acceptable(valid_proposal(), validation) is True


@pytest.mark.parametrize(
    ("proposal", "reason_fragment"),
    [
        (
            valid_proposal(
                corrected_zip_code="78209",
                information_added=True,
            ),
            "information was added",
        ),
        (valid_proposal(corrected_zip_code="7820"), "ZIP format is invalid"),
        (valid_proposal(corrected_city=""), "current City"),
        (valid_proposal(corrected_suite="C2"), "unsupported"),
    ],
)
def test_invalid_ai_corrections_are_rejected(proposal, reason_fragment) -> None:
    current = current_unresolved()
    if reason_fragment == "current City":
        current["City"] = "ALAMO HEIGHTS"
    validation = validate_ai_proposal(current["PropertyAddress"], current, proposal)
    assert validation.valid is False
    assert any(reason_fragment in reason for reason in validation.reasons)


def test_low_confidence_valid_suggestion_is_not_auto_accepted() -> None:
    proposal = valid_proposal(confidence=0.70)
    current = current_unresolved()
    validation = validate_ai_proposal(current["PropertyAddress"], current, proposal)
    assert validation.valid
    assert not is_auto_acceptable(proposal, validation)


@pytest.mark.parametrize(
    "fake",
    [FakeClient(output=None), FakeClient(error=TimeoutError("secret details"))],
)
def test_malformed_timeout_and_api_errors_are_safely_wrapped(fake) -> None:
    with pytest.raises(ReviewServiceError) as exc_info:
        review_address_with_openai(
            build_review_payload(current_unresolved()),
            reviewer_config(),
            client=fake,
        )
    assert "secret details" not in str(exc_info.value)


def test_confirmation_is_required_before_any_client_call() -> None:
    result = unresolved_result()
    client = FakeClient(output=valid_proposal())
    cache = {}
    run = review_eligible_rows(
        result,
        ReviewerConfig(api_key=""),
        cache,
        confirmed=False,
        client=client,
    )
    assert run.requested == 0
    assert client.responses.calls == []


def test_ai_mode_without_key_fails_safely_before_call() -> None:
    result = unresolved_result()
    client = FakeClient(output=valid_proposal())
    with pytest.raises(MissingAPIKeyError):
        review_eligible_rows(
            result,
            ReviewerConfig(api_key=""),
            {},
            confirmed=True,
            client=client,
        )
    assert client.responses.calls == []
    assert result.database.loc[1, "Parse Status"] == "Review Needed"


def test_only_review_needed_rows_are_sent_and_duplicate_click_is_cached() -> None:
    result = unresolved_result()
    client = FakeClient(output=valid_proposal())
    cache = {}
    config = reviewer_config(auto_accept=False)
    first = review_eligible_rows(
        result, config, cache, confirmed=True, client=client
    )
    second = review_eligible_rows(
        result, config, cache, confirmed=True, client=client
    )
    assert first.completed == 1
    assert second.requested == 0
    assert second.skipped_cached == 1
    assert len(client.responses.calls) == 1
    sent = client.responses.calls[0]["input"][1]["content"]
    assert "6104 BROADWAY" in sent
    assert "1 MAIN ST AUSTIN" not in sent
    assert result.database.loc[0, "Parse Status"] == "Parsed"


def test_auto_accept_can_be_enabled_but_is_off_by_default() -> None:
    result = unresolved_result()
    cache = {}
    client = FakeClient(output=valid_proposal())
    review_eligible_rows(
        result,
        reviewer_config(auto_accept=False),
        cache,
        confirmed=True,
        client=client,
    )
    assert result.database.loc[1, "Parse Status"] == "Review Needed"

    result = unresolved_result()
    run = review_eligible_rows(
        result,
        reviewer_config(auto_accept=True),
        {},
        confirmed=True,
        client=FakeClient(output=valid_proposal()),
    )
    assert run.auto_accepted == 1
    assert result.database.loc[1, "Parse Status"] == "Parsed"
    assert result.database.loc[1, "Suite"] == "C1"


def test_manual_acceptance_and_rejection_are_recorded() -> None:
    result = unresolved_result()
    cache = {}
    review_eligible_rows(
        result,
        reviewer_config(),
        cache,
        confirmed=True,
        client=FakeClient(output=valid_proposal()),
    )
    editor = build_review_editor(result, cache, "gpt-5-nano")
    editor.loc[0, "Accept Suggestion"] = True
    outcome = apply_review_editor(result, editor, cache)
    assert outcome.applied == 1
    assert result.database.loc[1, "Parse Status"] == "Parsed"
    assert bool(result.ai_review_log.iloc[0]["Correction Accepted"]) is True
    assert result.ai_review_log.iloc[0]["Final Review Status"] == "Accepted Manually"

    result = unresolved_result()
    cache = {}
    review_eligible_rows(
        result,
        reviewer_config(),
        cache,
        confirmed=True,
        client=FakeClient(output=valid_proposal()),
    )
    editor = build_review_editor(result, cache, "gpt-5-nano")
    editor.loc[0, "Reject Suggestion"] = True
    outcome = apply_review_editor(result, editor, cache)
    assert outcome.rejected == 1
    assert result.database.loc[1, "Parse Status"] == "Review Needed"
    assert result.ai_review_log.iloc[0]["Final Review Status"].startswith("Rejected")


def test_manual_edit_uses_same_traceability_validation() -> None:
    result = unresolved_result()
    editor = build_review_editor(result, {}, "gpt-5-nano")
    editor.loc[0, "Final Address"] = "6104 BROADWAY ST"
    editor.loc[0, "Final Suite"] = "C1"
    editor.loc[0, "Final City"] = "ALAMO HEIGHTS"
    outcome = apply_review_editor(result, editor, {})
    assert outcome.applied == 1
    assert result.database.loc[1, "Parse Status"] == "Parsed"
    assert result.ai_review_log.iloc[0]["Review Method"] == "Manual"

    result = unresolved_result()
    editor = build_review_editor(result, {}, "gpt-5-nano")
    editor.loc[0, "Final Address"] = "6104 BROADWAY ST"
    editor.loc[0, "Final Suite"] = "C1"
    editor.loc[0, "Final City"] = "ALAMO HEIGHTS"
    editor.loc[0, "Final Zip Code"] = "78209"
    outcome = apply_review_editor(result, editor, {})
    assert outcome.applied == 0
    assert outcome.errors
    assert result.database.loc[1, "Parse Status"] == "Review Needed"


def test_api_failure_preserves_python_result_and_excel_export() -> None:
    result = unresolved_result()
    original_row = result.database.loc[1].copy()
    run = review_eligible_rows(
        result,
        reviewer_config(),
        {},
        confirmed=True,
        client=FakeClient(error=TimeoutError("do not expose")),
    )
    assert run.failed == 1
    assert result.database.loc[1].equals(original_row)
    assert result.ai_review_log.iloc[0]["Validation Result"] == "API review failed"
    workbook = load_workbook(BytesIO(create_excel_workbook(result)))
    assert workbook["Database"].max_row == 3
    assert workbook["AI Review Log"].max_row == 2
    database_headers = [cell.value for cell in workbook["Database"][1]]
    assert "AI Explanation" not in database_headers
    assert "Parse Notes" not in database_headers
