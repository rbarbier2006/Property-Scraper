"""CSV cleaning, session data models, and in-memory Excel export."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
import re

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

from address_parser import parse_address_detailed


CANONICAL_COLUMNS = {
    "propertyid": "Property ID",
    "geographicid": "Geographic ID",
    "type": "Type",
    "propertyaddress": "PropertyAddress",
    "legaldescription": "Legal Description",
    "ownername": "Owner Name",
    "doingbusinessas": "Doing Business As",
    "appraisedvalue": "Appraised Value",
}

FINAL_PRIMARY_ORDER = [
    "Property ID",
    "Geographic ID",
    "Type",
    "PropertyAddress",
    "Address",
    "Suite",
    "City",
    "Zip Code",
    "Owner Name",
    "Doing Business As",
    "Appraised Value",
    "Parse Status",
]

GENERATED_OUTPUT_COLUMNS = ["Address", "Suite", "City", "Zip Code", "Parse Status"]
PROHIBITED_OUTPUT_HEADERS = {
    "parsenotes",
    "aiexplanation",
    "correctionnotes",
    "technicalerrordetails",
    "originalparsestatus",
    "reviewmethod",
    "aiconfidence",
    "aidecision",
    "correctionaccepted",
    "finalreviewstatus",
}

AI_REVIEW_LOG_COLUMNS = [
    "Original PropertyAddress",
    "Original Parse Status",
    "Python Address",
    "Python Suite",
    "Python City",
    "Python Zip Code",
    "Python Parse Status",
    "AI Proposed Address",
    "AI Proposed Suite",
    "AI Proposed City",
    "AI Proposed Zip Code",
    "AI Confidence",
    "AI Explanation",
    "AI Decision",
    "Information Added",
    "Manual Review Required",
    "Validation Result",
    "Correction Accepted",
    "Review Method",
    "Final Address",
    "Final Suite",
    "Final City",
    "Final Zip Code",
    "Final Parse Status",
    "Final Review Status",
]

INTERNAL_AI_LOG_COLUMNS = ["_Review Signature", "_Database Row"]

SUMMARY_LABELS = [
    ("total_uploaded", "Total uploaded rows"),
    ("personal_retained", "Personal rows retained"),
    ("real_removed", "Real rows removed"),
    ("blank_or_unexpected_excluded", "Blank or unexpected Type rows excluded"),
    ("first_pass_parsed", "Addresses parsed in first pass"),
    ("second_pass_corrected", "Addresses corrected in deterministic second pass"),
    ("addresses_parsed", "Addresses currently parsed"),
    ("addresses_review_needed", "Addresses still unresolved"),
    ("ai_eligible", "Addresses eligible for AI review"),
    ("ai_rows_reviewed", "Addresses reviewed by AI"),
    ("ai_corrections_accepted", "AI corrections accepted"),
    ("manual_corrections_accepted", "Manual corrections accepted"),
]


@dataclass(frozen=True)
class ColumnDetection:
    canonical_to_actual: dict[str, str]
    errors: list[str]

    @property
    def type_column(self) -> str | None:
        return self.canonical_to_actual.get("Type")

    @property
    def property_address_column(self) -> str | None:
        return self.canonical_to_actual.get("PropertyAddress")

    @property
    def legal_description_column(self) -> str | None:
        return self.canonical_to_actual.get("Legal Description")


@dataclass
class ProcessingResult:
    """All in-session processing state; only selected frames are exported."""

    database: pd.DataFrame
    review_needed: pd.DataFrame
    summary: dict[str, int]
    diagnostics: pd.DataFrame
    ai_review_log: pd.DataFrame


def normalize_header(header: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(header).strip().casefold())


def detect_columns(columns: list[object] | pd.Index) -> ColumnDetection:
    normalized_to_actual: dict[str, list[str]] = {}
    for column in columns:
        normalized_to_actual.setdefault(normalize_header(column), []).append(str(column))

    errors: list[str] = []
    for required_normalized, friendly_name in (
        ("type", "Type"),
        ("propertyaddress", "PropertyAddress or Property Address"),
    ):
        matches = normalized_to_actual.get(required_normalized, [])
        if not matches:
            errors.append(f"Missing required column: {friendly_name}.")
        elif len(matches) > 1:
            errors.append(
                f"Multiple columns match {friendly_name}: {', '.join(matches)}. "
                "Rename or remove the duplicate before processing."
            )

    canonical_to_actual: dict[str, str] = {}
    for normalized, canonical in CANONICAL_COLUMNS.items():
        matches = normalized_to_actual.get(normalized, [])
        if len(matches) == 1:
            canonical_to_actual[canonical] = matches[0]
    return ColumnDetection(canonical_to_actual=canonical_to_actual, errors=errors)


def read_csv_bytes(data: bytes) -> tuple[pd.DataFrame, str]:
    if not data:
        raise ValueError("The uploaded CSV is empty.")

    decode_errors: list[str] = []
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            frame = pd.read_csv(
                BytesIO(data),
                encoding=encoding,
                dtype=str,
                keep_default_na=False,
                na_filter=False,
            )
            return frame, encoding
        except UnicodeDecodeError as exc:
            decode_errors.append(f"{encoding}: {exc}")
        except pd.errors.EmptyDataError as exc:
            raise ValueError("The CSV has no header or data columns.") from exc
        except pd.errors.ParserError as exc:
            raise ValueError("The CSV structure could not be read.") from exc

    raise ValueError(
        "The CSV encoding could not be read as UTF-8, UTF-8 with BOM, or Latin-1. "
        + " | ".join(decode_errors)
    )


def _canonicalize_known_columns(frame: pd.DataFrame) -> pd.DataFrame:
    detection = detect_columns(frame.columns)
    if detection.errors:
        raise ValueError(" ".join(detection.errors))
    rename_map = {
        actual: canonical
        for canonical, actual in detection.canonical_to_actual.items()
        if actual != canonical
    }
    return frame.rename(columns=rename_map).copy()


def _empty_ai_log() -> pd.DataFrame:
    return pd.DataFrame(columns=INTERNAL_AI_LOG_COLUMNS + AI_REVIEW_LOG_COLUMNS)


def refresh_processing_result(result: ProcessingResult) -> None:
    """Recalculate derived views and metrics after AI or manual review."""

    result.review_needed = result.database.loc[
        result.database["Parse Status"].eq("Review Needed")
    ].copy()
    parsed_count = int(
        result.database["Parse Status"].isin(["Parsed", "AI Parsed"]).sum()
    )
    unresolved_count = int(result.database["Parse Status"].eq("Review Needed").sum())
    result.summary["addresses_parsed"] = parsed_count
    result.summary["addresses_review_needed"] = unresolved_count
    result.summary["ai_eligible"] = unresolved_count

    log = result.ai_review_log
    if log.empty:
        result.summary["ai_rows_reviewed"] = 0
        result.summary["ai_corrections_accepted"] = 0
        result.summary["manual_corrections_accepted"] = 0
        return

    ai_reviewed = log["Review Method"].eq("OpenAI")
    accepted = log["Correction Accepted"].map(
        lambda value: value is True or str(value).strip().casefold() == "true"
    )
    result.summary["ai_rows_reviewed"] = int(ai_reviewed.sum())
    result.summary["ai_corrections_accepted"] = int(
        (accepted & log["Review Method"].eq("OpenAI")).sum()
    )
    result.summary["manual_corrections_accepted"] = int(
        (accepted & log["Review Method"].eq("Manual")).sum()
    )


def process_dataframe(frame: pd.DataFrame) -> ProcessingResult:
    """Filter Personal rows and apply both deterministic parser stages."""

    working = _canonicalize_known_columns(frame)
    type_values = working["Type"].astype(str).str.strip().str.casefold()
    personal_mask = type_values.eq("personal")
    real_mask = type_values.eq("real")

    database = working.loc[personal_mask].copy().reset_index(drop=True)
    columns_to_drop = [
        column
        for column in database.columns
        if column == "Legal Description"
        or column in GENERATED_OUTPUT_COLUMNS
        or normalize_header(column) in PROHIBITED_OUTPUT_HEADERS
    ]
    if columns_to_drop:
        database = database.drop(columns=columns_to_drop)

    parsed_rows: list[dict[str, str]] = []
    diagnostic_rows: list[dict[str, object]] = []
    for database_row, value in enumerate(database["PropertyAddress"].tolist()):
        parsed, diagnostics = parse_address_detailed(value)
        parsed_rows.append({column: parsed[column] for column in GENERATED_OUTPUT_COLUMNS})
        diagnostic_rows.append(
            {
                "Database Row": database_row,
                "Parse Notes": diagnostics.notes,
                "Original Parse Status": diagnostics.original_status,
                "Review Method": diagnostics.review_method,
                "Second Pass Corrected": diagnostics.second_pass_corrected,
            }
        )

    parsed_frame = pd.DataFrame(parsed_rows, columns=GENERATED_OUTPUT_COLUMNS)
    for column in GENERATED_OUTPUT_COLUMNS:
        database[column] = parsed_frame[column] if not parsed_frame.empty else ""

    ordered_primary = [column for column in FINAL_PRIMARY_ORDER if column in database.columns]
    remaining = [column for column in database.columns if column not in ordered_primary]
    database = database.loc[:, ordered_primary + remaining]

    diagnostics = pd.DataFrame(
        diagnostic_rows,
        columns=[
            "Database Row",
            "Parse Notes",
            "Original Parse Status",
            "Review Method",
            "Second Pass Corrected",
        ],
    )
    first_pass_parsed = int(
        (diagnostics["Original Parse Status"] == "Parsed").sum()
    ) if not diagnostics.empty else 0
    second_pass_corrected = int(
        diagnostics["Second Pass Corrected"].fillna(False).sum()
    ) if not diagnostics.empty else 0
    summary = {
        "total_uploaded": int(len(working)),
        "personal_retained": int(personal_mask.sum()),
        "real_removed": int(real_mask.sum()),
        "blank_or_unexpected_excluded": int((~personal_mask & ~real_mask).sum()),
        "first_pass_parsed": first_pass_parsed,
        "second_pass_corrected": second_pass_corrected,
        "addresses_parsed": 0,
        "addresses_review_needed": 0,
        "ai_eligible": 0,
        "ai_rows_reviewed": 0,
        "ai_corrections_accepted": 0,
        "manual_corrections_accepted": 0,
    }
    result = ProcessingResult(
        database=database,
        review_needed=pd.DataFrame(columns=database.columns),
        summary=summary,
        diagnostics=diagnostics,
        ai_review_log=_empty_ai_log(),
    )
    refresh_processing_result(result)
    return result


def _summary_frame(summary: dict[str, int]) -> pd.DataFrame:
    return pd.DataFrame(
        [(label, summary.get(key, 0)) for key, label in SUMMARY_LABELS],
        columns=["Metric", "Count"],
    )


def _format_worksheet(worksheet, *, text_headers: set[str] | None = None) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    header_lookup = {cell.value: cell.column for cell in worksheet[1]}
    for header in text_headers or set():
        column_index = header_lookup.get(header)
        if column_index and worksheet.max_row > 1:
            for cells in worksheet.iter_cols(
                min_col=column_index,
                max_col=column_index,
                min_row=2,
                max_row=worksheet.max_row,
            ):
                for cell in cells:
                    cell.number_format = "@"

    for column_cells in worksheet.columns:
        header = str(column_cells[0].value or "")
        max_length = len(header)
        for cell in column_cells[1 : min(len(column_cells), 250)]:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        width_cap = 60 if "Address" in header or "Explanation" in header else 35
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
            max(max_length + 2, 12), width_cap
        )


def create_excel_workbook(result: ProcessingResult) -> bytes:
    """Build Database, Review Needed, Summary, and AI Review Log in memory."""

    refresh_processing_result(result)
    public_log = result.ai_review_log.reindex(columns=AI_REVIEW_LOG_COLUMNS).copy()
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        result.database.to_excel(writer, index=False, sheet_name="Database")
        result.review_needed.to_excel(writer, index=False, sheet_name="Review Needed")
        _summary_frame(result.summary).to_excel(
            writer, index=False, sheet_name="Processing Summary"
        )
        public_log.to_excel(writer, index=False, sheet_name="AI Review Log")

        workbook = writer.book
        text_headers = {
            header
            for header in result.database.columns
            if normalize_header(header).endswith("id")
            or "zipcode" in normalize_header(header)
        }
        _format_worksheet(workbook["Database"], text_headers=text_headers)
        _format_worksheet(workbook["Review Needed"], text_headers=text_headers)
        _format_worksheet(workbook["Processing Summary"])
        _format_worksheet(
            workbook["AI Review Log"],
            text_headers={
                "Python Zip Code",
                "AI Proposed Zip Code",
                "Final Zip Code",
            },
        )

    return output.getvalue()


__all__ = [
    "AI_REVIEW_LOG_COLUMNS",
    "ColumnDetection",
    "FINAL_PRIMARY_ORDER",
    "ProcessingResult",
    "create_excel_workbook",
    "detect_columns",
    "normalize_header",
    "process_dataframe",
    "read_csv_bytes",
    "refresh_processing_result",
]
